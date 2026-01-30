from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Q, Sum, Count
from decimal import Decimal
from django.utils import timezone
from datetime import datetime, timedelta

from .models import Plat, Panier, PanierItem, Commande, CommandeItem, Paiement, Caisse, Depense
from tables.models import TableRestaurant
from accounts.decorators import table_only, serveur_only, cuisinier_only, comptable_only
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
# from accounts.decorators import cuisinier_only
# from django.utils import timezone
# from datetime import datetime, timedelta
# from accounts.decorators import comptable_only

@login_required
def menu_list(request):
    """Liste des plats disponibles"""
    categorie = request.GET.get('categorie', '')
    recherche = request.GET.get('q', '')
    
    plats = Plat.objects.filter(disponible=True)
    
    if categorie:
        plats = plats.filter(categorie=categorie)
    
    if recherche:
        plats = plats.filter(
            Q(nom__icontains=recherche) | Q(description__icontains=recherche)
        )
    
    # Regrouper par catégorie
    plats_par_categorie = {}
    for categorie_code, categorie_nom in Plat.CATEGORIES:
        plats_categorie = plats.filter(categorie=categorie_code)
        if plats_categorie.exists():
            plats_par_categorie[categorie_nom] = plats_categorie
    
    # Récupérer le panier actif pour l'utilisateur table
    panier = None
    panier_count = 0
    if request.user.role == 'Rtable':
        try:
            table = request.user.table_restaurant
            panier = Panier.objects.filter(table=table, actif=True).first()
            if panier:
                panier_count = panier.items.count()
        except:
            pass
    
    context = {
        'plats_par_categorie': plats_par_categorie,
        'categories': Plat.CATEGORIES,
        'panier_count': panier_count,
    }
    
    return render(request, 'menu/menu_list.html', context)


@login_required
def plat_list_api(request):
    """API pour récupérer les plats (format JSON)"""
    plats = Plat.objects.filter(disponible=True).values(
        'id', 'nom', 'description', 'prix', 'categorie', 'image'
    )
    return JsonResponse(list(plats), safe=False)


@login_required
@table_only
def panier_view(request):
    """Afficher le panier"""
    try:
        table = request.user.table_restaurant
        panier = Panier.objects.filter(table=table, actif=True).first()
        
        if not panier:
            panier = Panier.objects.create(table=table)
        
        items = panier.items.select_related('plat').all()
        total = panier.total
        
        context = {
            'panier': panier,
            'items': items,
            'total': total,
        }
        
        return render(request, 'menu/panier.html', context)
    except:
        messages.error(request, "Erreur lors de l'accès au panier.")
        return redirect('menu:menu_list')


@login_required
@table_only
def ajouter_au_panier(request, plat_id):
    """Ajouter un plat au panier"""
    if request.method == 'POST':
        try:
            plat = get_object_or_404(Plat, id=plat_id, disponible=True)
            table = request.user.table_restaurant
            
            # Récupérer ou créer le panier actif
            panier, created = Panier.objects.get_or_create(
                table=table,
                actif=True
            )
            
            # Récupérer la quantité
            quantite = int(request.POST.get('quantite', 1))
            
            # Validation de la quantité
            if quantite < 1 or quantite > 10:
                messages.error(request, "La quantité doit être entre 1 et 10.")
                return redirect('menu:menu_list')
            
            # Vérifier si le plat est déjà dans le panier
            item, item_created = PanierItem.objects.get_or_create(
                panier=panier,
                plat=plat,
                defaults={
                    'quantite': quantite,
                    'prix_unitaire': plat.prix
                }
            )
            
            if not item_created:
                # Mettre à jour la quantité
                nouvelle_quantite = item.quantite + quantite
                if nouvelle_quantite > 10:
                    messages.error(request, "Vous ne pouvez pas commander plus de 10 unités d'un même plat.")
                    return redirect('menu:panier')
                
                item.quantite = nouvelle_quantite
                item.save()
                messages.success(request, f"{plat.nom} - quantité mise à jour dans le panier.")
            else:
                messages.success(request, f"{plat.nom} ajouté au panier.")
            
            return redirect('menu:panier')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout au panier : {str(e)}")
            return redirect('menu:menu_list')
    
    return redirect('menu:menu_list')


@login_required
@table_only
def modifier_panier_item(request, item_id):
    """Modifier la quantité d'un article dans le panier"""
    if request.method == 'POST':
        try:
            table = request.user.table_restaurant
            panier = get_object_or_404(Panier, table=table, actif=True)
            item = get_object_or_404(PanierItem, id=item_id, panier=panier)
            
            quantite = int(request.POST.get('quantite', 1))
            
            if quantite < 1 or quantite > 10:
                messages.error(request, "La quantité doit être entre 1 et 10.")
                return redirect('menu:panier')
            
            item.quantite = quantite
            item.save()
            
            messages.success(request, "Quantité mise à jour.")
            return redirect('menu:panier')
            
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
            return redirect('menu:panier')
    
    return redirect('menu:panier')


@login_required
@table_only
def supprimer_panier_item(request, item_id):
    """Supprimer un article du panier"""
    try:
        table = request.user.table_restaurant
        panier = get_object_or_404(Panier, table=table, actif=True)
        item = get_object_or_404(PanierItem, id=item_id, panier=panier)
        
        item.delete()
        messages.success(request, "Article supprimé du panier.")
        
    except Exception as e:
        messages.error(request, f"Erreur : {str(e)}")
    
    return redirect('menu:panier')


@login_required
@table_only
def vider_panier(request):
    """Vider le panier"""
    try:
        table = request.user.table_restaurant
        panier = get_object_or_404(Panier, table=table, actif=True)
        
        panier.items.all().delete()
        messages.success(request, "Panier vidé.")
        
    except Exception as e:
        messages.error(request, f"Erreur : {str(e)}")
    
    return redirect('menu:panier')


@login_required
@table_only
@transaction.atomic
def valider_commande(request):
    """Valider le panier et créer une commande"""
    if request.method == 'POST':
        try:
            table = request.user.table_restaurant
            panier = get_object_or_404(Panier, table=table, actif=True)
            
            # Vérifier que le panier n'est pas vide
            if not panier.items.exists():
                messages.error(request, "Votre panier est vide.")
                return redirect('menu:panier')
            
            # Calculer le montant total
            montant_total = panier.total
            
            # Créer la commande
            commande = Commande.objects.create(
                table=table,
                montant_total=montant_total,
                statut='en_attente'
            )
            
            # Créer les items de la commande
            for item in panier.items.all():
                CommandeItem.objects.create(
                    commande=commande,
                    plat=item.plat,
                    quantite=item.quantite,
                    prix_unitaire=item.prix_unitaire
                )
            
            # Désactiver le panier
            panier.actif = False
            panier.save()
            
            # Mettre à jour l'état de la table
            table.etat = 'attente'
            table.save()
            
            messages.success(request, f"Commande #{commande.id} créée avec succès ! Montant total : {montant_total} GNF")
            return redirect('menu:mes_commandes')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la validation de la commande : {str(e)}")
            return redirect('menu:panier')
    
    return redirect('menu:panier')


@login_required
@table_only
def mes_commandes(request):
    """Liste des commandes de la table"""
    try:
        table = request.user.table_restaurant
        commandes = Commande.objects.filter(table=table).order_by('-date_commande')
        
        context = {
            'commandes': commandes,
        }
        
        return render(request, 'menu/mes_commandes.html', context)
    except:
        messages.error(request, "Erreur lors de la récupération des commandes.")
        return redirect('menu:menu_list')


@login_required
def detail_commande(request, commande_id):
    """Détail d'une commande"""
    commande = get_object_or_404(Commande, id=commande_id)
    
    # Vérifier les permissions
    if request.user.role == 'Rtable':
        if commande.table != request.user.table_restaurant:
            messages.error(request, "Vous n'avez pas accès à cette commande.")
            return redirect('menu:mes_commandes')
    elif request.user.role not in ['Rservent', 'Radmin', 'Rcomptable']:
        messages.error(request, "Vous n'avez pas accès à cette page.")
        return redirect('menu:menu_list')
    
    items = commande.items.select_related('plat').all()
    
    context = {
        'commande': commande,
        'items': items,
    }
    
    return render(request, 'menu/detail_commande.html', context)
@login_required
@cuisinier_only
def tableau_bord_cuisinier(request):
    """Tableau de bord pour le cuisinier"""
    # Commandes en préparation
    commandes_en_preparation = Commande.objects.filter(
        statut__in=['en_attente', 'en_preparation']
    ).select_related('table').prefetch_related('items__plat').order_by('date_commande')
    
    # Statistiques
    stats = {
        'total_plats': Plat.objects.count(),
        'plats_disponibles': Plat.objects.filter(disponible=True).count(),
        'plats_indisponibles': Plat.objects.filter(disponible=False).count(),
        'commandes_en_attente': commandes_en_preparation.count(),
    }
    
    context = {
        'commandes': commandes_en_preparation,
        'stats': stats,
    }
    
    return render(request, 'menu/cuisinier_dashboard.html', context)


@login_required
@cuisinier_only
def gestion_plats_cuisinier(request):
    """Gestion des plats par le cuisinier"""
    categorie = request.GET.get('categorie', '')
    recherche = request.GET.get('q', '')
    
    plats = Plat.objects.all()
    
    if categorie:
        plats = plats.filter(categorie=categorie)
    
    if recherche:
        plats = plats.filter(
            Q(nom__icontains=recherche) | Q(description__icontains=recherche)
        )
    
    # Regrouper par catégorie
    plats_par_categorie = {}
    for categorie_code, categorie_nom in Plat.CATEGORIES:
        plats_categorie = plats.filter(categorie=categorie_code)
        if plats_categorie.exists():
            plats_par_categorie[categorie_nom] = plats_categorie
    
    context = {
        'plats_par_categorie': plats_par_categorie,
        'categories': Plat.CATEGORIES,
    }
    
    return render(request, 'menu/cuisinier_plats.html', context)


@login_required
@cuisinier_only
def ajouter_plat(request):
    """Ajouter un nouveau plat"""
    if request.method == 'POST':
        try:
            nom = request.POST.get('nom')
            description = request.POST.get('description', '')
            prix = request.POST.get('prix')
            categorie = request.POST.get('categorie')
            image = request.FILES.get('image')
            disponible = request.POST.get('disponible') == 'on'
            
            # Validation
            if not nom or not prix or not categorie:
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return redirect('menu:ajouter_plat')
            
            # Créer le plat
            plat = Plat.objects.create(
                nom=nom,
                description=description,
                prix=prix,
                categorie=categorie,
                disponible=disponible,
                image=image
            )
            
            messages.success(request, f"Plat '{nom}' ajouté avec succès!")
            return redirect('menu:gestion_plats_cuisinier')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout du plat: {str(e)}")
            return redirect('menu:ajouter_plat')
    
    context = {
        'categories': Plat.CATEGORIES,
    }
    return render(request, 'menu/cuisinier_ajouter_plat.html', context)


@login_required
@cuisinier_only
def modifier_plat(request, plat_id):
    """Modifier un plat existant"""
    plat = get_object_or_404(Plat, id=plat_id)
    
    if request.method == 'POST':
        try:
            plat.nom = request.POST.get('nom')
            plat.description = request.POST.get('description', '')
            plat.prix = request.POST.get('prix')
            plat.categorie = request.POST.get('categorie')
            plat.disponible = request.POST.get('disponible') == 'on'
            
            # Mise à jour de l'image si fournie
            if 'image' in request.FILES:
                plat.image = request.FILES['image']
            
            plat.save()
            
            messages.success(request, f"Plat '{plat.nom}' modifié avec succès!")
            return redirect('menu:gestion_plats_cuisinier')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification: {str(e)}")
    
    context = {
        'plat': plat,
        'categories': Plat.CATEGORIES,
    }
    return render(request, 'menu/cuisinier_modifier_plat.html', context)


@login_required
@cuisinier_only
def changer_disponibilite_plat(request, plat_id):
    """Activer/Désactiver la disponibilité d'un plat"""
    if request.method == 'POST':
        plat = get_object_or_404(Plat, id=plat_id)
        plat.disponible = not plat.disponible
        plat.save()
        
        statut = "disponible" if plat.disponible else "indisponible"
        messages.success(request, f"Plat '{plat.nom}' marqué comme {statut}.")
    
    return redirect('menu:gestion_plats_cuisinier')


@login_required
@cuisinier_only
def changer_statut_commande(request, commande_id):
    """Changer le statut d'une commande (cuisinier)"""
    if request.method == 'POST':
        commande = get_object_or_404(Commande, id=commande_id)
        action = request.POST.get('action')
        
        try:
            if action == 'demarrer':
                if commande.statut == 'en_attente':
                    commande.statut = 'en_preparation'
                    commande.save()
                    messages.success(request, f"Commande #{commande.id} en préparation.")
                else:
                    messages.error(request, "Cette commande n'est pas en attente.")
            
            elif action == 'terminer':
                if commande.statut == 'en_preparation':
                    commande.statut = 'prete'
                    commande.save()
                    
                    # Mettre à jour l'état de la table
                    table = commande.table
                    table.etat = 'servie'
                    table.save()
                    
                    messages.success(request, f"Commande #{commande.id} prête à servir!")
                else:
                    messages.error(request, "Cette commande n'est pas en préparation.")
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    return redirect('menu:tableau_bord_cuisinier')
@login_required
@comptable_only
def tableau_bord_comptable(request):
    """Tableau de bord pour le comptable"""
    # Récupérer ou créer la caisse
    caisse, created = Caisse.objects.get_or_create(id=1)
    
    # Statistiques du jour
    aujourd_hui = timezone.now().date()
    debut_jour = timezone.make_aware(datetime.combine(aujourd_hui, datetime.min.time()))
    fin_jour = timezone.make_aware(datetime.combine(aujourd_hui, datetime.max.time()))
    
    # Paiements du jour
    paiements_jour = Paiement.objects.filter(
        date_paiement__range=(debut_jour, fin_jour)
    )
    total_recettes_jour = paiements_jour.aggregate(
        total=Sum('montant')
    )['total'] or Decimal('0.00')
    
    # Dépenses du jour
    depenses_jour = Depense.objects.filter(
        date__range=(debut_jour, fin_jour)
    )
    total_depenses_jour = depenses_jour.aggregate(
        total=Sum('montant')
    )['total'] or Decimal('0.00')
    
    # Commandes par statut
    commandes_stats = {
        'payees': Commande.objects.filter(statut='payee').count(),
        'en_attente': Commande.objects.filter(statut='en_attente').count(),
        'servies': Commande.objects.filter(statut='servie').count(),
    }
    
    # Dernières transactions
    dernieres_recettes = paiements_jour.select_related('commande__table').order_by('-date_paiement')[:5]
    dernieres_depenses = depenses_jour.select_related('enregistre_par').order_by('-date')[:5]
    
    context = {
        'caisse': caisse,
        'total_recettes_jour': total_recettes_jour,
        'total_depenses_jour': total_depenses_jour,
        'benefice_jour': total_recettes_jour - total_depenses_jour,
        'commandes_stats': commandes_stats,
        'dernieres_recettes': dernieres_recettes,
        'dernieres_depenses': dernieres_depenses,
        'nombre_paiements_jour': paiements_jour.count(),
        'nombre_depenses_jour': depenses_jour.count(),
    }
    
    return render(request, 'menu/comptable_dashboard.html', context)


@login_required
@comptable_only
def liste_paiements(request):
    """Liste de tous les paiements"""
    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    paiements = Paiement.objects.select_related('commande__table').all()
    
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
            paiements = paiements.filter(date_paiement__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
            date_fin_obj = timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
            paiements = paiements.filter(date_paiement__lte=date_fin_obj)
        except ValueError:
            pass
    
    paiements = paiements.order_by('-date_paiement')
    
    # Total
    total_paiements = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    
    context = {
        'paiements': paiements,
        'total_paiements': total_paiements,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'menu/comptable_paiements.html', context)


@login_required
@comptable_only
def liste_depenses(request):
    """Liste de toutes les dépenses"""
    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    depenses = Depense.objects.select_related('enregistre_par').all()
    
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
            depenses = depenses.filter(date__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
            date_fin_obj = timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
            depenses = depenses.filter(date__lte=date_fin_obj)
        except ValueError:
            pass
    
    depenses = depenses.order_by('-date')
    
    # Total
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    
    context = {
        'depenses': depenses,
        'total_depenses': total_depenses,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'menu/comptable_depenses.html', context)


@login_required
@comptable_only
@transaction.atomic
def ajouter_depense(request):
    """Enregistrer une nouvelle dépense"""
    if request.method == 'POST':
        try:
            motif = request.POST.get('motif')
            montant = Decimal(request.POST.get('montant'))
            
            # Validation
            if not motif or montant <= 0:
                messages.error(request, "Veuillez remplir tous les champs correctement.")
                return redirect('menu:ajouter_depense')
            
            # Vérifier le solde de la caisse
            caisse, created = Caisse.objects.get_or_create(id=1)
            
            if caisse.solde_actuel < montant:
                messages.error(request, 
                    f"Solde insuffisant! Solde actuel: {caisse.solde_actuel} GNF, "
                    f"Montant demandé: {montant} GNF"
                )
                return redirect('menu:ajouter_depense')
            
            # Créer la dépense
            depense = Depense.objects.create(
                motif=motif,
                montant=montant,
                enregistre_par=request.user
            )
            
            # Déduire du solde de la caisse
            caisse.solde_actuel -= montant
            caisse.save()
            
            messages.success(request, 
                f"Dépense de {montant} GNF enregistrée avec succès! "
                f"Nouveau solde: {caisse.solde_actuel} GNF"
            )
            return redirect('menu:liste_depenses')
            
        except ValueError:
            messages.error(request, "Montant invalide.")
            return redirect('menu:ajouter_depense')
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
            return redirect('menu:ajouter_depense')
    
    # Récupérer le solde actuel
    caisse, created = Caisse.objects.get_or_create(id=1)
    
    context = {
        'solde_actuel': caisse.solde_actuel,
    }
    
    return render(request, 'menu/comptable_ajouter_depense.html', context)


@login_required
@comptable_only
def rapport_financier(request):
    """Générer un rapport financier"""
    # Période par défaut: mois en cours
    aujourd_hui = timezone.now().date()
    debut_mois = aujourd_hui.replace(day=1)
    
    date_debut = request.GET.get('date_debut', debut_mois.strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', aujourd_hui.strftime('%Y-%m-%d'))
    
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
        date_fin_obj = timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
        date_debut_obj = timezone.make_aware(datetime.combine(date_debut_obj, datetime.min.time()))
    except ValueError:
        messages.error(request, "Format de date invalide.")
        return redirect('menu:tableau_bord_comptable')
    
    # Paiements de la période
    paiements = Paiement.objects.filter(
        date_paiement__range=(date_debut_obj, date_fin_obj)
    )
    total_recettes = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    
    # Dépenses de la période
    depenses = Depense.objects.filter(
        date__range=(date_debut_obj, date_fin_obj)
    )
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    
    # Bénéfice
    benefice = total_recettes - total_depenses
    
    # Récupérer la caisse
    caisse, created = Caisse.objects.get_or_create(id=1)
    
    context = {
        'date_debut': date_debut,
        'date_fin': date_fin,
        'total_recettes': total_recettes,
        'total_depenses': total_depenses,
        'benefice': benefice,
        'nombre_paiements': paiements.count(),
        'nombre_depenses': depenses.count(),
        'solde_caisse': caisse.solde_actuel,
        'paiements': paiements.select_related('commande__table').order_by('-date_paiement'),
        'depenses': depenses.select_related('enregistre_par').order_by('-date'),
    }
    
    return render(request, 'menu/comptable_rapport.html', context)



@login_required
@comptable_only
def export_rapport_excel(request):
    """Exporter le rapport financier en Excel"""
    # Récupérer les paramètres de date
    date_debut = request.GET.get('date_debut', timezone.now().date().replace(day=1).strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', timezone.now().date().strftime('%Y-%m-%d'))
    
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
        date_fin_obj = timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
        date_debut_obj = timezone.make_aware(datetime.combine(date_debut_obj, datetime.min.time()))
    except ValueError:
        messages.error(request, "Format de date invalide.")
        return redirect('menu:tableau_bord_comptable')
    
    # Récupérer les données
    paiements = Paiement.objects.filter(date_paiement__range=(date_debut_obj, date_fin_obj))
    depenses = Depense.objects.filter(date__range=(date_debut_obj, date_fin_obj))
    
    total_recettes = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    benefice = total_recettes - total_depenses
    
    # Créer le classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Financier"
    
    # Style d'en-tête
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Titre
    ws['A1'] = 'RAPPORT FINANCIER - RESTAURO'
    ws['A1'].font = Font(bold=True, size=16, color="4F46E5")
    ws.merge_cells('A1:D1')
    
    # Période
    ws['A2'] = f'Période : {date_debut} au {date_fin}'
    ws['A2'].font = Font(size=11)
    ws.merge_cells('A2:D2')
    
    # Résumé
    ws['A4'] = 'RÉSUMÉ'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws.merge_cells('A4:B4')
    
    ws['A5'] = 'Total Recettes'
    ws['B5'] = f'{total_recettes} GNF'
    ws['A6'] = 'Total Dépenses'
    ws['B6'] = f'{total_depenses} GNF'
    ws['A7'] = 'Bénéfice'
    ws['B7'] = f'{benefice} GNF'
    ws['B7'].font = Font(bold=True, color="10B981" if benefice >= 0 else "EF4444")
    
    # Liste des paiements
    ws['A9'] = 'PAIEMENTS'
    ws['A9'].font = header_font
    ws['A9'].fill = header_fill
    ws.merge_cells('A9:D9')
    
    ws['A10'] = 'Date'
    ws['B10'] = 'Commande #'
    ws['C10'] = 'Table'
    ws['D10'] = 'Montant (GNF)'
    for cell in ['A10', 'B10', 'C10', 'D10']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
    
    row = 11
    for paiement in paiements:
        ws[f'A{row}'] = paiement.date_paiement.strftime('%d/%m/%Y %H:%M')
        ws[f'B{row}'] = paiement.commande.id
        ws[f'C{row}'] = paiement.commande.table.numero_table
        ws[f'D{row}'] = float(paiement.montant)
        row += 1
    
    # Liste des dépenses
    row += 2
    ws[f'A{row}'] = 'DÉPENSES'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Date'
    ws[f'B{row}'] = 'Motif'
    ws[f'C{row}'] = 'Enregistré par'
    ws[f'D{row}'] = 'Montant (GNF)'
    for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
    
    row += 1
    for depense in depenses:
        ws[f'A{row}'] = depense.date.strftime('%d/%m/%Y %H:%M')
        ws[f'B{row}'] = depense.motif
        ws[f'C{row}'] = depense.enregistre_par.login if depense.enregistre_par else 'N/A'
        ws[f'D{row}'] = float(depense.montant)
        row += 1
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Préparer la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rapport_financier_{date_debut}_{date_fin}.xlsx'
    wb.save(response)
    
    return response


@login_required
@comptable_only
def export_rapport_pdf(request):
    """Exporter le rapport financier en PDF"""
    # Récupérer les paramètres de date
    date_debut = request.GET.get('date_debut', timezone.now().date().replace(day=1).strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', timezone.now().date().strftime('%Y-%m-%d'))
    
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
        date_fin_obj = timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
        date_debut_obj = timezone.make_aware(datetime.combine(date_debut_obj, datetime.min.time()))
    except ValueError:
        messages.error(request, "Format de date invalide.")
        return redirect('menu:tableau_bord_comptable')
    
    # Récupérer les données
    paiements = Paiement.objects.filter(date_paiement__range=(date_debut_obj, date_fin_obj))
    depenses = Depense.objects.filter(date__range=(date_debut_obj, date_fin_obj))
    
    total_recettes = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    benefice = total_recettes - total_depenses
    
    # Créer le PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = styles['Title']
    elements.append(Paragraph('RAPPORT FINANCIER - RESTAURO', title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Période
    elements.append(Paragraph(f'Période : {date_debut} au {date_fin}', styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Résumé
    resume_data = [
        ['RÉSUMÉ', ''],
        ['Total Recettes', f'{total_recettes} GNF'],
        ['Total Dépenses', f'{total_depenses} GNF'],
        ['Bénéfice', f'{benefice} GNF'],
    ]
    
    resume_table = Table(resume_data, colWidths=[3*inch, 3*inch])
    resume_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(resume_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Paiements
    elements.append(Paragraph('PAIEMENTS', styles['Heading2']))
    paiements_data = [['Date', 'Commande', 'Table', 'Montant']]
    for p in paiements:
        paiements_data.append([
            p.date_paiement.strftime('%d/%m/%Y'),
            f'#{p.commande.id}',
            p.commande.table.numero_table,
            f'{p.montant} GNF'
        ])
    
    paiements_table = Table(paiements_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2*inch])
    paiements_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(paiements_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Dépenses
    elements.append(Paragraph('DÉPENSES', styles['Heading2']))
    depenses_data = [['Date', 'Motif', 'Enregistré par', 'Montant']]
    for d in depenses:
        depenses_data.append([
            d.date.strftime('%d/%m/%Y'),
            d.motif[:30],
            d.enregistre_par.login if d.enregistre_par else 'N/A',
            f'{d.montant} GNF'
        ])
    
    depenses_table = Table(depenses_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch])
    depenses_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(depenses_table)
    
    # Générer le PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=rapport_financier_{date_debut}_{date_fin}.pdf'
    
    return response
